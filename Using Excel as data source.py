import openpyxl

path="E:\data\data.xlsx"
workbook=openpyxl.load_workbook(path)

sheet=workbook.active
rows=sheet.max_row
col=sheet.max_column

print(rows)
print(col)
for r in range(1,rows+1):
	for c in range(1,col+1):
		print(sheet.cell(row=r,coloumn=c).value,end="    ")

	print()