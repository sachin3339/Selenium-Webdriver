import openpyxl
from selenium import webdriver

driver = webdriver.Firefox(executable_path="C://Selenium//driver//chromedriver_win32//geckodriver.exe")
URL = "http://www.google.com"
driver.get("http://localhost:3000/login")

def getRowCount(file,sheetName):
	workbook=openpyxl.load_workbook(file)
	sheet=workbook.get_Sheet_by_name(sheetName)
	return(sheet.max_row)	


def readData(file,sheetName,rownum,coloumnno):
	workbook=openpyxl.load_workbook(file)
	sheet=workbook.get_Sheet_by_name(sheetName)
	return sheet.cell(row=r,coloumn=c).value



path="E:\data\data.xlsx"
workbook=openpyxl.load_workbook(path)
rows=getRowCount(path,'Sheet1')


for r in range(2,3):
	username=readData(path,"Sheet1",r,1)
	password=readData(path,"Sheet1",r,1)
	print(username)
	print(password)